   
import openpyxl

# Spreadsheet

def get_credentials(user_type):
    # Load Excel file
    workbook = openpyxl.load_workbook("testdata.xlsx")
    sheet = workbook["Credentials"]  # Sheet name

    # Loop through rows and match user_type
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_type:  # Check user type in column A
            username = row[1]    # Column B
            password = row[2]    # Column C
            print("Username:", username)
            print("Password:", password)
            return username, password  # Return values

    return None, None  # If user_type not found




def get_register_data(user_type):
    workbook = openpyxl.load_workbook("testdata.xlsx")
    sheet = workbook["Register"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_type:
            data = {
                "name": row[1],
                "email": row[2],
                "password": row[3],
                "gender": row[4],
                "day": row[5],
                "month": row[6],
                "year": row[7]
            }
            return data

    return None